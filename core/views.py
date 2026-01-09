import logging
import io
import tempfile

from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import (
    DetailView, 
    CreateView, 
    DeleteView,
    ListView, 
    View, 
    TemplateView
)
from django.shortcuts import get_object_or_404
from django.db.models import Count
from .forms import (
    AddCarForm, 
    OwnerForm, 
    ServiceForm, 
    OutlayFrom, 
    CarServiceForm,
    InvoiceUploadForm,
    InvoiceItemForm
)
from .models import (
    Car, 
    Owner, 
    CarPhoto, 
    Service, 
    Outlay, 
    CarServiceState,
    ServiceEventSchema,
    Invoice,
    InvoiceItem,
    OutlayTypeChoice,
    OutlayCategoryChoice,
    Notifications
)
from .services import create_outlay
from django.utils import timezone
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.core.exceptions import ValidationError
from django.conf import settings
from django.db import transaction
from pathlib import Path
import os
from decimal import Decimal
from .services import (
    create_car_with_photos, 
    update_car_with_photos, 
    delete_car, 
    get_outlay, 
    create_outlay, 
    get_outlay_form_data, 
    update_outlay,
    save_or_update_car_service_state,
    create_service_events_from_services,
    PDFCore,
    decode_unicode_escapes,
    create_car_service_plan,
)
from .constants import DEFAULT_SERVICE_SCHEMA

logger = logging.getLogger(__name__)

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cars_active = Car.objects.filter(status="Active").select_related('owner')
        cars_pending = Car.objects.filter(status="Await").select_related('owner')
        context['cars_active'] = cars_active
        context['cars_pending'] = cars_pending
        context['form'] = AddCarForm()
        return context
    

class AddCarView(LoginRequiredMixin, View):
    template_name = "car/create.html"
    
    def get(self, request):
        form = AddCarForm()
        owners = Owner.objects.all()
        return render(request, self.template_name, {
            'form': form,
            'owners': owners,
            'owner_form': OwnerForm()
        })
    
    def post(self, request):
        logger.info(f"Request method: {request.method}")

        result = create_car_with_photos(
            form_data=request.POST,
            files=request.FILES,
            form_class=AddCarForm,
        )

        if result["success"]:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse(
                    {
                        "status": "ok",
                        "id": str(result["car"].uuid),
                        "message": result["message"],
                    }
                )
            return redirect('car-detail', pk=result["car"].uuid)
        
        status_code = 500 if "__all__" in result["errors"] else 400
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse(
                {
                    "status": "error",
                    "errors": result["errors"],
                },
                status=status_code,
            )
        form = AddCarForm(request.POST)
        owners = Owner.objects.all()
        return render(request, self.template_name, {
            'form': form,
            'owners': owners,
            'owner_form': OwnerForm(),
            'errors': result["errors"],
        })


class CarDetailView(LoginRequiredMixin, DetailView):
    model = Car
    template_name = "car/detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = AddCarForm(instance=self.object)
        context["photos"] = CarPhoto.objects.filter(car=self.object)
        # Get outlays for this car
        outlays = Outlay.objects.filter(cars=self.object).select_related('amount').order_by('-created_at', '-updated_at')
        context["outlays"] = outlays
        
        # Calculate total outlay amount
        total = 0
        for outlay in outlays:
            if outlay.amount.full_price:
                total += float(outlay.amount.full_price)
            elif outlay.amount.price_per_item and outlay.amount.item_count:
                total += float(outlay.amount.price_per_item) * int(outlay.amount.item_count)
        context["outlays_total"] = total
        
        # Get car service state if exists
        try:
            context["car_service_state"] = CarServiceState.objects.get(car=self.object)
        except CarServiceState.DoesNotExist:
            context["car_service_state"] = None
        
        return context


class CarDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        logger.info(f"Request method: {request.method}, UUID: {pk}")

        result = delete_car(car_uuid=str(pk))

        if result["success"]:
            return JsonResponse(
                {
                    "status": "ok",
                    "message": result["message"],
                }
            )
        status_code = 404 if "not found" in result["message"].lower() else 500
        return JsonResponse(
            {
                "status": "error",
                "errors": result["errors"],
            },
            status=status_code,
        )

class CarSaveView(LoginRequiredMixin, View):
    pass

class CarUnsaveView(LoginRequiredMixin, View):
    pass

class CarSaveListView(LoginRequiredMixin, ListView):
    model = Car
    template_name = "car/save_list.html"

    def get_queryset(self):
        user=self.request.user
        cars= user.cars.filter(is_saved=True).order_by("-created_at")
        return cars

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = AddCarForm()

class CarUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        logger.info(f"Request method: {request.method}, UUID: {pk}")

        result = update_car_with_photos(
            car_uuid=str(pk),
            form_data=request.POST,
            files=request.FILES,
            form_class=AddCarForm,
        )

        if result["success"]:
            return JsonResponse(
                {
                    "status": "ok",
                    "id": str(result["car"].uuid),
                    "message": result["message"],
                }
            )
        status_code = (
            404
            if "not found" in result["message"].lower()
            else (500 if "__all__" in result["errors"] else 400)
        )
        return JsonResponse(
            {
                "status": "error",
                "errors": result["errors"],
            },
            status=status_code,
        )


class OwnerListView(LoginRequiredMixin, ListView):
    model = Owner
    paginate_by = 20
    template_name = "owner/list.html"

    def get_queryset(self):
        return Owner.objects.annotate(cars_count=Count("cars")).order_by(
            "first_name", "last_name"
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = OwnerForm()
        return context


class OwnerCreateView(LoginRequiredMixin, View):
    def post(self, request):
        if request.content_type == 'application/json':
            import json
            data = json.loads(request.body)
            form = OwnerForm(data)
        else:
            form = OwnerForm(request.POST)

        if form.is_valid():
            owner = form.save()
            return JsonResponse(
                {
                    "status": "ok",
                    "id": str(owner.uuid),
                    "message": "Owner successfully added!",
                }
            )

        return JsonResponse(
            {
                "status": "error",
                "errors": form.errors,
            },
            status=400,
        )


class OwnerUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            owner = Owner.objects.get(uuid=pk)
        except Owner.DoesNotExist:
            return JsonResponse(
                {
                    "status": "error",
                    "errors": {"__all__": ["Owner not found"]},
                },
                status=404,
            )

        form = OwnerForm(request.POST, instance=owner)

        if form.is_valid():
            owner = form.save()
            return JsonResponse({
                "status": "ok",
                "id": str(owner.uuid),
                "message": "Owner successfully updated!"
            })

        return JsonResponse({
            "status": "error",
            "errors": form.errors
        }, status=400)


class OwnerDetailView(LoginRequiredMixin, DetailView):
    model = Owner
    template_name = "owner/detail.html"


class OwnerDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            owner = Owner.objects.get(uuid=pk)
            owner.delete()
            return JsonResponse({
                "status": "ok",
                "message": "Owner successfully deleted!",
            })
        except Owner.DoesNotExist:
            return JsonResponse({
                "status": "error",
                "errors": {"__all__": ["Owner not found"]},
            }, status=404)


class ServiceListView(LoginRequiredMixin, ListView):
    model = Service
    template_name = "service/list.html"
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ServiceForm()
        return context


class ServiceCreateView(LoginRequiredMixin, View):
    def post(self, request):
        form = ServiceForm(request.POST)
        if form.is_valid():
            service = form.save()
            return JsonResponse({
                "status": "ok",
                "id": str(service.uuid),
                "message": "Service successfully added!",
            })
        return JsonResponse({
            "status": "error",
            "errors": form.errors,
        }, status=400)


class ServiceUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            service = Service.objects.get(uuid=pk)
        except Service.DoesNotExist:
            return JsonResponse({
                "status": "error",
                "errors": {"__all__": ["Service not found"]},
            }, status=404)
        
        form = ServiceForm(request.POST, instance=service)
        if form.is_valid():
            service = form.save()
            return JsonResponse({
                "status": "ok",
                "id": str(service.uuid),
                "message": "Service successfully updated!",
            })
        return JsonResponse({
            "status": "error",
            "errors": form.errors,
        }, status=400)


class ServiceDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            service = Service.objects.get(uuid=pk)
            service.delete()
            return JsonResponse({
                "status": "ok",
                "message": "Service successfully deleted!",
            })
        except Service.DoesNotExist:
            return JsonResponse({
                "status": "error",
                "errors": {"__all__": ["Service not found"]},
            }, status=404)


class ServiceDetailView(LoginRequiredMixin, DetailView):
    model = Service
    template_name = "service/detail.html"


class OutlayView(LoginRequiredMixin, View):
    template_name = "outlay.html"

    def get(self, request):
        # Get filter parameter
        filter_type = request.GET.get('type', 'all')  # all, service, other
        filter_category = request.GET.get('category', 'all')  # all, fuel, parts, documents, another
        
        # Get all outlays
        outlays_qs = Outlay.objects.select_related('amount').prefetch_related('cars').order_by('-created_at', '-updated_at')
        
        # Apply type filter
        if filter_type == 'service':
            outlays_qs = outlays_qs.filter(type='service')
        elif filter_type == 'other':
            outlays_qs = outlays_qs.filter(type='other')
        
        # Apply category filter
        if filter_category != 'all':
            outlays_qs = outlays_qs.filter(category=filter_category)
        
        # Convert to list of dicts for template compatibility
        outlays = []
        for outlay in outlays_qs:
            cars_list = list(outlay.cars.all())
            if cars_list:
                for car in cars_list:
                    outlays.append({
                        'uuid': outlay.uuid,
                        'type': outlay.type,
                        'category': outlay.category,
                        'category_name': outlay.category_name,
                        'service_name': outlay.service_name,
                        'name': outlay.name,
                        'comment': outlay.comment,
                        'description': outlay.description,  # For backward compatibility
                        'created_at': outlay.created_at,
                        'updated_at': outlay.updated_at,
                        'cars__mark': car.mark,
                        'cars__model': car.model,
                        'cars__license_plate': car.license_plate,
                        'amount__price_per_item': outlay.amount.price_per_item,
                        'amount__item_count': outlay.amount.item_count,
                        'amount__full_price': outlay.amount.full_price,
                    })
            else:
                # If no cars associated, still show the outlay
                outlays.append({
                    'uuid': outlay.uuid,
                    'type': outlay.type,
                    'category': outlay.category,
                    'category_name': outlay.category_name,
                    'service_name': outlay.service_name,
                    'name': outlay.name,
                    'comment': outlay.comment,
                    'description': outlay.description,  # For backward compatibility
                    'created_at': outlay.created_at,
                    'updated_at': outlay.updated_at,
                    'cars__mark': '',
                    'cars__model': '',
                    'cars__license_plate': '',
                    'amount__price_per_item': outlay.amount.price_per_item,
                    'amount__item_count': outlay.amount.item_count,
                    'amount__full_price': outlay.amount.full_price,
                })
        
        context = {
            "outlays": outlays,
            "form": OutlayFrom(),
            "filter_type": filter_type,
            "filter_category": filter_category,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        form = OutlayFrom(request.POST)
        
        # Check if car field has value even if form is not valid
        car_value = request.POST.get('car')
        if not car_value:
            form.add_error('car', 'Оберіть автомобіль')
        
        if not form.is_valid():
            # If form is invalid, render with errors
            filter_type = request.GET.get('type', 'all')
            filter_category = request.GET.get('category', 'all')
            outlays_qs = Outlay.objects.select_related('amount').prefetch_related('cars').order_by('-created_at', '-updated_at')
            
            if filter_type == 'service':
                outlays_qs = outlays_qs.filter(type='service')
            elif filter_type == 'other':
                outlays_qs = outlays_qs.filter(type='other')
            
            if filter_category != 'all':
                outlays_qs = outlays_qs.filter(category=filter_category)
            
            outlays = []
            for outlay in outlays_qs:
                cars_list = list(outlay.cars.all())
                if cars_list:
                    for car in cars_list:
                        outlays.append({
                            'uuid': outlay.uuid,
                            'type': outlay.type,
                            'category': outlay.category,
                            'category_name': outlay.category_name,
                            'service_name': outlay.service_name,
                            'name': outlay.name,
                            'comment': outlay.comment,
                            'description': outlay.description,
                            'created_at': outlay.created_at,
                            'updated_at': outlay.updated_at,
                            'cars__mark': car.mark,
                            'cars__model': car.model,
                            'cars__license_plate': car.license_plate,
                            'amount__price_per_item': outlay.amount.price_per_item,
                            'amount__item_count': outlay.amount.item_count,
                            'amount__full_price': outlay.amount.full_price,
                        })
            
            context = {
                "outlays": outlays,
                "form": form,
                "filter_type": filter_type,
                "filter_category": filter_category,
            }
            return render(request, self.template_name, context)
        
        cd = form.cleaned_data
        
        # Validate name for service type
        if cd['service_type'] == 'service' and not cd.get('name'):
            form.add_error('name', 'Назва витрати обов\'язкова для типу "Сервіс"')
            filter_type = request.GET.get('type', 'all')
            filter_category = request.GET.get('category', 'all')
            outlays_qs = Outlay.objects.select_related('amount').prefetch_related('cars').order_by('-created_at', '-updated_at')
            
            if filter_type == 'service':
                outlays_qs = outlays_qs.filter(type='service')
            elif filter_type == 'other':
                outlays_qs = outlays_qs.filter(type='other')
            
            if filter_category != 'all':
                outlays_qs = outlays_qs.filter(category=filter_category)
            
            outlays = []
            for outlay in outlays_qs:
                cars_list = list(outlay.cars.all())
                if cars_list:
                    for car in cars_list:
                        outlays.append({
                            'uuid': outlay.uuid,
                            'type': outlay.type,
                            'category': outlay.category,
                            'category_name': outlay.category_name,
                            'service_name': outlay.service_name,
                            'name': outlay.name,
                            'comment': outlay.comment,
                            'description': outlay.description,
                            'created_at': outlay.created_at,
                            'updated_at': outlay.updated_at,
                            'cars__mark': car.mark,
                            'cars__model': car.model,
                            'cars__license_plate': car.license_plate,
                            'amount__price_per_item': outlay.amount.price_per_item,
                            'amount__item_count': outlay.amount.item_count,
                            'amount__full_price': outlay.amount.full_price,
                        })
            
            context = {
                "outlays": outlays,
                "form": form,
                "filter_type": filter_type,
                "filter_category": filter_category,
            }
            return render(request, self.template_name, context)
        
        # Create outlay using service function
        outlay = create_outlay(
            type=cd['service_type'],
            name=cd.get('name', ''),
            car=cd['car'],  # Single car instead of list
            price_per_item=cd.get('price_per_item') or 0,
            item_count=cd.get('item_count') or 0,
            created_at=cd['date'],
            full_price=cd.get('full_price'),
            category=cd.get('category'),
            category_name=cd.get('category_name'),
            service_name=cd.get('service_name'),
            comment=cd.get('comment'),
            description=cd.get('description'),  # For backward compatibility
        )
        
        return redirect('outlay')


class OutlatDetailView(LoginRequiredMixin, View):
    template_name = "outlay_detail.html"
    view_template_name = "outlay_detail_view.html"

    def get(self, request, pk):
        outlay = get_outlay(pk)
        
        # Try to find related invoice from comment
        invoice = None
        if outlay.comment and "Автоматично створено з фактури:" in outlay.comment:
            import re
            match = re.search(r'Автоматично створено з фактури:\s*([^,]+)', outlay.comment)
            if match:
                invoice_name = match.group(1).strip()
                try:
                    invoice = Invoice.objects.filter(name=invoice_name).first()
                except Exception as e:
                    logger.exception(f"Error finding invoice for outlay {pk}: {e}")
        
        # Handle invoice PDF download
        if request.GET.get('download_invoice') == 'true' and invoice:
            from django.http import Http404, HttpResponse
            from django.conf import settings
            import os
            
            file_path = Path(settings.MEDIA_ROOT) / invoice.file_path
            if not file_path.exists():
                raise Http404("PDF файл не знайдено")
            
            filename = os.path.basename(invoice.file_path)
            
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            response = HttpResponse(file_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        # Check if edit mode is requested
        edit_mode = request.GET.get('edit', 'false').lower() == 'true'
        
        if edit_mode:
            # Show edit form
            form = get_outlay_form_data(pk)
            context = {
                "outlay": outlay,
                "outlay_uuid": pk,
                "form": form,
                "invoice": invoice,
            }
            return render(request, self.template_name, context)
        else:
            # Show view mode
            context = {
                "outlay": outlay,
                "outlay_uuid": pk,
                "invoice": invoice,
            }
            return render(request, self.view_template_name, context)

    def post(self, request, pk):
        try:
            outlay = get_outlay(pk)
        except Outlay.DoesNotExist:
            return redirect('outlay')
        
        form = OutlayFrom(request.POST)
        
        # Check if car field has value even if form is not valid
        car_value = request.POST.get('car')
        if not car_value:
            form.add_error('car', 'Оберіть автомобіль')
        
        if form.is_valid():
            cd = form.cleaned_data
            
            # Validate name for service type
            if cd['service_type'] == 'service' and not cd.get('name'):
                form.add_error('name', 'Назва витрати обов\'язкова для типу "Сервіс"')
                context = {
                    "outlay": outlay,
                    "outlay_uuid": pk,
                    "form": form,
                }
                return render(request, self.template_name, context)
            
            # Validate car is selected
            if not cd.get('car'):
                form.add_error('car', 'Оберіть автомобіль')
                context = {
                    "outlay": outlay,
                    "outlay_uuid": pk,
                    "form": form,
                }
                return render(request, self.template_name, context)
            
            try:
                # Update outlay using service function
                update_outlay(pk, form)
                
                # Get car UUID for redirect
                updated_outlay = get_outlay(pk)
                cars = list(updated_outlay.cars.all())
                car_uuid = cars[0].uuid if cars else None
                
                # Check if request is from car detail page or car outlays page
                referer = request.META.get('HTTP_REFERER', '')
                if car_uuid and '/core/cars/' in referer:
                    if '/outlays/' in referer:
                        return redirect('car-outlays', pk=car_uuid)
                    return redirect('car-detail', pk=car_uuid)
                
                # Redirect to view mode after successful update
                return redirect('outlay_detail', pk=pk)
            except Exception as e:
                logger.error(f"Error updating outlay {pk}: {str(e)}")
                form.add_error(None, f'Помилка оновлення витрати: {str(e)}')
                context = {
                    "outlay": outlay,
                    "outlay_uuid": pk,
                    "form": form,
                }
                return render(request, self.template_name, context)
        else:
            # Form is invalid, render with errors
            context = {
                "outlay": outlay,
                "outlay_uuid": pk,
                "form": form,
            }
            return render(request, self.template_name, context)


class OutlayDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            outlay = Outlay.objects.get(uuid=pk)
            car_uuid = None
            if hasattr(outlay, 'car') and outlay.car:
                car_uuid = outlay.car.uuid
            else:
                cars = list(outlay.cars.all())
                car_uuid = cars[0].uuid if cars else None
            
            outlay.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "ok",
                    "message": "Витрату успішно видалено",
                })
            
            referer = request.META.get('HTTP_REFERER', '')
            if car_uuid or '/core/cars/' in referer:
                if not car_uuid:
                    import re
                    match = re.search(r'/core/cars/([0-9a-f-]+)', referer)
                    if match:
                        car_uuid = match.group(1)
                
                if car_uuid:
                    return redirect('car-detail', pk=car_uuid)
            
            return redirect('outlay')
        except Outlay.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "error",
                    "errors": {"__all__": ["Витрату не знайдено"]},
                }, status=404)
            return redirect('outlay')
        except Exception as e:
            logger.error(f"Error deleting outlay {pk}: {str(e)}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "error",
                    "errors": {"__all__": [f"Помилка видалення: {str(e)}"]},
                }, status=500)
            return redirect('outlay')

class CarServiceCreate(LoginRequiredMixin, CreateView):
    model = CarServiceState
    form_class = CarServiceForm
    template_name = "car_service_plan/create.html"
    success_url = "/"

    def get(self, request):
        form = CarServiceForm()
        cars = Car.objects.all()
        cars_mileage = {str(car.uuid): car.mileage for car in cars}
        
        # Використовуємо дефолтну схему з constants.py, якщо немає в БД
        default_schema = ServiceEventSchema.objects.filter(
            is_default=True
        ).first()
        
        schema_data = default_schema.schema if default_schema else DEFAULT_SERVICE_SCHEMA
        
        return render(request, self.template_name, {
            'form': form,
            'cars': cars,
            'cars_mileage': cars_mileage,
            'default_schema': schema_data
        })

    def post(self, request, *args, **kwargs):
        """Перевизначити post для обробки AJAX запитів"""
        try:
            form = self.form_class(request.POST)
            if form.is_valid():
                return self.form_valid(form)
            else:
                return self.form_invalid(form)
        except Exception as e:
            # Обробка будь-яких несподіваних помилок
            logger.exception(f"Error in CarServiceCreate.post: {e}")
            return JsonResponse({
                "status": "error",
                "errors": {"__all__": [f"Помилка сервера: {str(e)}"]},
            }, status=500)
    
    def form_valid(self, form):
        try:
            car = form.cleaned_data['car']
            service_plan = form.cleaned_data['service_plan']
            services_json = form.cleaned_data.get('services_json')
            mileage = form.cleaned_data.get('mileage') or car.mileage
            
            old_mileage = car.mileage
            
            mileage_updated = False
            if mileage != car.mileage:
                car.mileage = mileage
                car.save(update_fields=['mileage'])
                mileage_updated = True
            
            try:
                car_service_state = save_or_update_car_service_state(
                    car=car,
                    service_plan=service_plan,
                    mileage=mileage
                )
            except (ValueError, Exception) as exc:
                logger.exception(f"Error in save_or_update_car_service_state: {exc}")
                return JsonResponse({
                    "status": "error",
                    "errors": {"__all__": [str(exc)]},
                }, status=400)

            # Створюємо ServiceEvent записи безпосередньо зі списку сервісів (без парсингу JSON)
            if services_json:
                try:
                    created_events = create_service_events_from_services(
                        car=car,
                        services=services_json,
                        mileage=mileage
                    )
                except (ValueError, Exception) as exc:
                    logger.exception(f"Error in create_service_events_from_services: {exc}")
                    return JsonResponse({
                        "status": "error",
                        "errors": {"__all__": [str(exc)]},
                    }, status=400)

            from django.urls import reverse
            
            response_data = {
                "status": "ok",
                "schema": car_service_state.service_plan,
                "car_id": str(car.uuid),
                "message": "Service plan schema calculated successfully!",
                "redirect_url": reverse("car-detail", kwargs={"pk": car.uuid}),
            }
            
            if mileage_updated:
                response_data["warning"] = f"Пробіг автомобіля оновлено з {old_mileage} км на {mileage} км в базі даних."
            
            return JsonResponse(response_data)
        except Exception as e:
            # Обробка будь-яких несподіваних помилок
            logger.exception(f"Error in CarServiceCreate.form_valid: {e}")
            return JsonResponse({
                "status": "error",
                "errors": {"__all__": [f"Помилка сервера: {str(e)}"]},
            }, status=500)
    
    def form_invalid(self, form):
        return JsonResponse({
            "status": "error",
            "errors": form.errors,
        }, status=400)



class CarServiceDeleteView(LoginRequiredMixin, View):
    def post(self, request, car_pk):
        try:
            car_service_state = get_object_or_404(CarServiceState, car__pk=car_pk)
            car_service_state.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "success",
                    "message": "Схему сервісного плану успішно видалено"
                })
            
            return redirect('car-detail', pk=car_pk)
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "error",
                    "message": str(e)
                }, status=400)
            raise


class CarServiceDetailView(LoginRequiredMixin, DetailView):
    model = CarServiceState
    template_name = "car_service_plan/detail.html"
    context_object_name = "car_service_plan"

    def get_object(self):
        car_pk = self.kwargs.get('car_pk')
        return get_object_or_404(CarServiceState, car__pk=car_pk)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        car_service_plan = self.object
        
        # Get current mileage from the car
        current_mileage = car_service_plan.car.mileage
        
        # Calculate overdue for each service
        services = car_service_plan.service_plan.get('services', [])
        for service in services:
            next_service = service.get('next_service')
            if next_service and current_mileage > next_service:
                service['overdue_km'] = current_mileage - next_service
            else:
                service['overdue_km'] = None
        
        context['current_mileage'] = current_mileage
        return context


class CarServiceUpdateView(LoginRequiredMixin, View):
    """View для оновлення конкретного сервісу - встановлює last_service_km на поточний пробіг"""
    
    @transaction.atomic
    def post(self, request, car_pk, service_key):
        try:
            car_service_state = CarServiceState.objects.select_for_update().get(car__pk=car_pk)
            car = car_service_state.car
            current_mileage = car.mileage
            
            # Знайти сервіс за key
            services = car_service_state.service_plan.get('services', [])
            service_found = False
            
            for service in services:
                if service.get('key') == service_key:
                    # Оновити last_service_km на поточний пробіг
                    service['last_service_km'] = current_mileage
                    service_found = True
                    break
            
            if not service_found:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Сервіс не знайдено'
                }, status=404)
            
            # Перерахувати всі сервіси з новими даними
            updated_services = create_car_service_plan(
                plan_schema={'services': services},
                current_mileage=current_mileage
            )
            
            # Оновити service_plan
            car_service_state.service_plan['services'] = updated_services
            car_service_state.save(update_fields=['service_plan', 'updated_at'])
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'message': 'Сервіс успішно оновлено'
                })
            
            return redirect('car-service-plan-detail', car_pk=car_pk)
            
        except CarServiceState.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'message': 'Схема сервісного плану не знайдена'
                }, status=404)
            return redirect('car-detail', pk=car_pk)
        except Exception as e:
            logger.exception(f"Error updating service: {e}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'message': f'Помилка оновлення: {str(e)}'
                }, status=500)
            return redirect('car-service-plan-detail', car_pk=car_pk)


class CarOutlaysView(LoginRequiredMixin, DetailView):
    model = Car
    template_name = "car/outlays.html"
    context_object_name = "car"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get outlays for this car
        outlays = Outlay.objects.filter(cars=self.object).select_related('amount').order_by('-created_at', '-updated_at')
        context["outlays"] = outlays
        
        # Calculate total outlay amount
        total = 0
        for outlay in outlays:
            if outlay.amount.full_price:
                total += float(outlay.amount.full_price)
            elif outlay.amount.price_per_item and outlay.amount.item_count:
                total += float(outlay.amount.price_per_item) * int(outlay.amount.item_count)
        context["outlays_total"] = total
        
        return context


class CarOutlaysExportView(LoginRequiredMixin, View):
    """Export car outlays to Excel for Financial Director and Accountant"""
    
    def get(self, request, pk):
        try:
            car = Car.objects.get(uuid=pk)
        except Car.DoesNotExist:
            return redirect('cars')
        
        # Get outlays for this car
        outlays = Outlay.objects.filter(cars=car).select_related('amount').order_by('created_at')
        
        # Create Excel file
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            from datetime import datetime
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Витрати"
            
            # Header style
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers for Financial Director and Accountant
            headers = [
                "Дата",
                "Тип витрати",
                "Категорія",
                "Назва витрати",
                "Сервіс",
                "Автомобіль",
                "VIN код",
                "Номерний знак",
                "Кількість",
                "Ціна за одиницю (PLN)",
                "Загальна сума (PLN)",
                "ПДВ %",
                "Сума ПДВ (PLN)",
                "Коментар",
                "Створено",
                "Оновлено"
            ]
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style
            
            # Write data
            for row_num, outlay in enumerate(outlays, 2):
                car_obj = outlay.cars.first()
                
                # Calculate values from OutlayAmount structure
                # Priority: use what's stored, calculate if needed
                price_per_item = float(outlay.amount.price_per_item) if outlay.amount.price_per_item else None
                item_count = float(outlay.amount.item_count) if outlay.amount.item_count else None
                total_price = float(outlay.amount.full_price) if outlay.amount.full_price else None
                
                # If total_price is missing but we have price_per_item and item_count, calculate it
                if not total_price and price_per_item and item_count:
                    total_price = price_per_item * item_count
                
                # If price_per_item is missing but we have total_price and item_count, calculate it
                if not price_per_item and total_price and item_count and item_count > 0:
                    price_per_item = total_price / item_count
                
                # If item_count is missing but we have total_price and price_per_item, calculate it
                if not item_count and total_price and price_per_item and price_per_item > 0:
                    item_count = total_price / price_per_item
                
                # Extract tax info if available (from comment or other sources)
                tax_percent = None
                tax_price = None
                if outlay.comment:
                    import re
                    tax_match = re.search(r'ПДВ[:\s]+(\d+(?:\.\d+)?)', outlay.comment, re.IGNORECASE)
                    if tax_match:
                        tax_percent = float(tax_match.group(1))
                
                # Get display values
                type_display = dict(OutlayTypeChoice.choices).get(outlay.type, outlay.type)
                category_display = ""
                if outlay.category:
                    category_display = dict(OutlayCategoryChoice.choices).get(outlay.category, outlay.category)
                elif outlay.category_name:
                    category_display = outlay.category_name
                
                row_data = [
                    outlay.created_at.strftime("%d.%m.%Y") if outlay.created_at else "",
                    type_display,
                    category_display,
                    outlay.name or "",
                    outlay.service_name or "",
                    f"{car_obj.mark} {car_obj.model} {car_obj.year}" if car_obj else "",
                    car_obj.vin_code if car_obj else "",
                    car_obj.license_plate if car_obj else "",
                    item_count if item_count else "",
                    price_per_item if price_per_item else "",
                    total_price if total_price else "",
                    tax_percent if tax_percent else "",
                    tax_price if tax_price else "",
                    outlay.comment or "",
                    outlay.created_at.strftime("%d.%m.%Y %H:%M") if outlay.created_at else "",
                    outlay.updated_at.strftime("%d.%m.%Y %H:%M") if outlay.updated_at else "",
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = border_style
                    if col_num in [9, 10, 11, 12, 13]:  # Numeric columns
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(header)
                for row in ws[column_letter]:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Freeze first row
            ws.freeze_panes = 'A2'
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"Витрати_{car.mark}_{car.model}_{car.year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except ImportError:
            # If openpyxl is not installed, return error message
            from django.contrib import messages
            messages.error(request, 'Бібліотека openpyxl не встановлена. Встановіть її для експорту Excel.')
            return redirect('car-outlays', pk=pk)
        except Exception as e:
            logger.exception(f"Error exporting outlays to Excel: {e}")
            from django.contrib import messages
            messages.error(request, f'Помилка експорту: {str(e)}')
            return redirect('car-outlays', pk=pk)


class ServiceEventSchemaDefaultView(LoginRequiredMixin, TemplateView):
    template_name = "service_event_schema/default_schema.html"
    
    def get_context_data(self, **kwargs):
        
        context = super().get_context_data(**kwargs)
        default_schema = ServiceEventSchema.objects.filter(
            is_default=True
        ).first()
        
        if not default_schema:
            default_schema = ServiceEventSchema.objects.create(
                schema_name="Дефолтна схема обслуговування",
                schema=DEFAULT_SERVICE_SCHEMA,
                is_default=True
            )
        
        context['service_event_schema'] = default_schema
        return context


class InvoiceListView(LoginRequiredMixin, ListView):
    model = Invoice
    template_name = "invoice/list.html"
    context_object_name = "invoices"
    paginate_by = 20

    def get_queryset(self):
        return Invoice.objects.prefetch_related('items').order_by('-created_at')


class InvoiceUploadView(LoginRequiredMixin, View):
    template_name = "invoice/upload.html"

    def get(self, request):
        return render(request, self.template_name, {
            "form": InvoiceUploadForm()
        })

    def post(self, request):
        form = InvoiceUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form})

        pdf_file = form.cleaned_data["pdf_file"]
        name = form.cleaned_data.get("name") or pdf_file.name

        # ---------- save file ----------
        upload_dir = Path(settings.MEDIA_ROOT) / "invoices"
        upload_dir.mkdir(parents=True, exist_ok=True)

        file_path = upload_dir / pdf_file.name
        with open(file_path, "wb+") as f:
            for chunk in pdf_file.chunks():
                f.write(chunk)

        # ---------- parse ----------
        try:
            parser = PDFCore(file_path)
            parsed_data = parser.parse()

            table_data = parsed_data.get("table", [])
            if not table_data:
                logger.warning("No table data found", extra={"parsed_data": parsed_data})
                form.add_error(
                    "pdf_file",
                    "Не вдалося знайти позиції у фактурі. Формат PDF не підтримується."
                )
                return render(request, self.template_name, {"form": form})

            # ---------- create invoice ----------
            invoice = Invoice.objects.create(
                name=name,
                file_path=str(file_path.relative_to(settings.MEDIA_ROOT)),
                invoice_data=parsed_data,
            )

            total_amount = Decimal("0")

            def to_decimal(val, default="0"):
                if not val:
                    return Decimal(default)
                if isinstance(val, (int, float)):
                    return Decimal(str(val))
                return Decimal(str(val).replace(" ", "").replace(",", "."))

            # ---------- create items ----------
            validation_errors_summary = {}
            SERVICE_GAS_UUID = "63d70638-32be-4959-8496-598a0c651f9d"
            car_matches = {}  # Store car matches for UI display
            
            # Get ServiceGas
            try:
                service_gas = Service.objects.get(uuid=SERVICE_GAS_UUID)
            except Service.DoesNotExist:
                logger.warning(f"ServiceGas with UUID {SERVICE_GAS_UUID} not found")
                service_gas = None
            
            def find_car_by_vin_or_plate(vin_or_plate):
                """Find car by VIN code or license plate"""
                if not vin_or_plate:
                    return None
                
                vin_or_plate = vin_or_plate.strip()
                if not vin_or_plate:
                    return None
                
                # Try to find by VIN code first
                car = Car.objects.filter(vin_code__iexact=vin_or_plate).first()
                if car:
                    return car
                
                # Try to find by license plate
                car = Car.objects.filter(license_plate__iexact=vin_or_plate).first()
                if car:
                    return car
                
                return None
            
            for row in table_data:
                try:
                    price_netto = to_decimal(row.get("price_netto"))
                    tax_price = to_decimal(row.get("tax_price"))
                    price_brutto = to_decimal(row.get("price_brutto"))
                    
                    # Clean item_name from Unicode escape sequences if any
                    item_name = row.get("item_name", "")
                    if item_name:
                        # Decode any remaining Unicode escapes
                        item_name = item_name.encode().decode('unicode_escape') if '\\u' in item_name else item_name
                    
                    current_car_vin = row.get("current_car_vin")
                    
                    item = InvoiceItem.objects.create(
                        invoice=invoice,
                        item_id=str(row.get("id")),
                        item_name=item_name[:1000] if item_name else "",  # Ensure max length
                        amount=to_decimal(row.get("amount", "1")),
                        price_netto=price_netto,
                        tax_percent=to_decimal(row.get("tax_percent", "23")),
                        tax_price=tax_price,
                        price_brutto=price_brutto,
                        current_car_vin=current_car_vin,
                    )
                    
                    # Try to find car and create outlay
                    car = None
                    if current_car_vin:
                        car = find_car_by_vin_or_plate(current_car_vin)
                        if car:
                            car_matches[str(item.uuid)] = {
                                'car_uuid': str(car.uuid),
                                'car_name': f"{car.mark} {car.model} {car.year}",
                                'vin_or_plate': current_car_vin
                            }
                            
                            # Create outlay for this item
                            if service_gas:
                                try:
                                    outlay = create_outlay(
                                        type='service',
                                        name=item_name[:255] if item_name else f"Витрата з фактури {invoice.name}",
                                        car=car,
                                        price_per_item=float(price_netto),
                                        item_count=float(item.amount),
                                        created_at=invoice.created_at or timezone.now(),
                                        full_price=float(price_brutto),
                                        service_name=service_gas.name,
                                        comment=f"Автоматично створено з фактури: {invoice.name}, позиція: {item.item_id}",
                                    )
                                    logger.info(f"Created outlay {outlay.uuid} for car {car.uuid} from invoice item {item.uuid}")
                                except Exception as e:
                                    logger.exception(f"Error creating outlay for invoice item {item.uuid}: {e}")
                        else:
                            # Car not found
                            car_matches[str(item.uuid)] = {
                                'car_found': False,
                                'vin_or_plate': current_car_vin
                            }
                    
                    # Store validation errors if any
                    if row.get("validation_errors"):
                        validation_errors_summary[str(item.uuid)] = row["validation_errors"]

                    total_amount += price_brutto

                except Exception as e:
                    logger.exception("Error creating invoice item", extra={"row": row})
                    continue

            invoice.invoice_amount = total_amount
            invoice.save(update_fields=["invoice_amount"])
            
            # Store validation errors, total validation, and car matches in invoice_data
            if validation_errors_summary or parsed_data.get("total_validation_error") or car_matches:
                if not invoice.invoice_data:
                    invoice.invoice_data = {}
                if validation_errors_summary:
                    invoice.invoice_data["validation_errors"] = validation_errors_summary
                if parsed_data.get("total_validation_error"):
                    invoice.invoice_data["total_validation_error"] = parsed_data["total_validation_error"]
                if car_matches:
                    invoice.invoice_data["car_matches"] = car_matches
                invoice.save(update_fields=["invoice_data"])

            return redirect("invoice-detail", pk=invoice.uuid)

        except Exception as e:
            logger.exception("Error parsing invoice PDF")
            form.add_error("pdf_file", f"Помилка обробки PDF: {e}")
            return render(request, self.template_name, {"form": form})


class NotificationsView(LoginRequiredMixin, TemplateView):
    template_name = "notifications/list.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Try to get real car data, fallback to mock if not found
        try:
            car = Car.objects.get(uuid='acb3a277-7f5a-4fb1-876b-7a147d9fe64e')
            default_car_uuid = str(car.uuid)
            default_license_plate = car.license_plate or 'WPI0316N'
        except Car.DoesNotExist:
            default_car_uuid = 'acb3a277-7f5a-4fb1-876b-7a147d9fe64e'
            default_license_plate = 'WPI0316N'
        
        # Get some real cars for variety
        real_cars = list(Car.objects.all()[:5])
        car_index = 0
        
        # Helper function to get car UUID as string
        def get_car_uuid_str(index):
            if real_cars and len(real_cars) > index:
                return str(real_cars[index % len(real_cars)].uuid)
            return default_car_uuid
        
        # Helper function to get car license plate
        def get_car_license_plate(index):
            if real_cars and len(real_cars) > index:
                car = real_cars[index % len(real_cars)]
                return car.license_plate if car.license_plate else default_license_plate
            return default_license_plate
        
        # Helper function to get overdue service info for a car
        def get_overdue_service_info(car_uuid_str):
            try:
                car = Car.objects.get(uuid=car_uuid_str)
                current_mileage = car.mileage
                
                try:
                    car_service_state = CarServiceState.objects.get(car=car)
                    services = car_service_state.service_plan.get('services', [])
                    
                    # Find overdue services
                    overdue_services = []
                    for service in services:
                        next_service = service.get('next_service')
                        if next_service and current_mileage > next_service:
                            overdue_km = current_mileage - next_service
                            overdue_services.append({
                                'name': service.get('name', 'Невідомий сервіс'),
                                'overdue_km': overdue_km,
                                'next_service': next_service,
                                'current_mileage': current_mileage,
                            })
                    
                    if overdue_services:
                        # Return the most overdue service
                        most_overdue = max(overdue_services, key=lambda x: x['overdue_km'])
                        return most_overdue
                except CarServiceState.DoesNotExist:
                    pass
            except Car.DoesNotExist:
                pass
            return None
        
        # Mock data for notifications (fake data as requested)
        mock_notifications = [
            {
                'message': 'Користувачу Іван Петренко надіслано прохання оновити свій пробіг',
                'message_type': 'mileage_update_request',
                'date': '11.01.2026',
                'time': '14:30',
                'icon': 'info',
                'color': '#2563eb',
                'car_uuid': get_car_uuid_str(0),
                'car_license_plate': get_car_license_plate(0),
            },
            {
                'message': 'Користувач Олександр Коваленко успішно оновив пробіг',
                'message_type': 'mileage_updated',
                'date': '11.01.2026',
                'time': '13:15',
                'icon': 'success',
                'color': '#10b981',
                'car_uuid': get_car_uuid_str(1),
                'car_license_plate': get_car_license_plate(1),
            },
        ]
        
        # Add service warning notifications with real overdue data
        service_warning_cars = [
            (default_car_uuid, default_license_plate, 'Марія Сидоренко', '10.01.2026', '16:45', 'масла'),
            (get_car_uuid_str(2), get_car_license_plate(2), 'Дмитро Мельник', '10.01.2026', '11:20', 'фільтра повітря'),
            (default_car_uuid, default_license_plate, 'Андрій Бондаренко', '09.01.2026', '10:15', 'фільтра повітря'),
        ]
        
        for car_uuid, license_plate, user_name, date, time, service_name in service_warning_cars:
            overdue_info = get_overdue_service_info(car_uuid)
            if overdue_info:
                message = f'Користувачу {user_name} надіслано попередження про скору заміну {service_name}'
                message += f' ({overdue_info["name"]})'
            else:
                message = f'Користувачу {user_name} надіслано попередження про скору заміну {service_name}'
            
            notification = {
                'message': message,
                'message_type': 'service_warning',
                'date': date,
                'time': time,
                'icon': 'warning',
                'color': '#f59e0b',
                'car_uuid': car_uuid,
                'car_license_plate': license_plate,
            }
            
            if overdue_info:
                notification['overdue_service'] = overdue_info
            
            mock_notifications.append(notification)
        
        # Add more mock notifications
        mock_notifications.extend([
            {
                'message': 'Користувач Наталія Шевченко успішно оновив пробіг',
                'message_type': 'mileage_updated',
                'date': '09.01.2026',
                'time': '15:30',
                'icon': 'success',
                'color': '#10b981',
                'car_uuid': get_car_uuid_str(3),
                'car_license_plate': get_car_license_plate(3),
            },
            {
                'message': 'Користувач Сергій Ткаченко успішно оновив пробіг',
                'message_type': 'mileage_updated',
                'date': '08.01.2026',
                'time': '09:45',
                'icon': 'success',
                'color': '#10b981',
                'car_uuid': get_car_uuid_str(4),
                'car_license_plate': get_car_license_plate(4),
            },
            {
                'message': 'Користувачу Олена Лисенко надіслано прохання оновити свій пробіг',
                'message_type': 'mileage_update_request',
                'date': '08.01.2026',
                'time': '14:00',
                'icon': 'info',
                'color': '#2563eb',
                'car_uuid': default_car_uuid,
                'car_license_plate': default_license_plate,
            },
        ])
        
        context['notifications'] = mock_notifications
        return context


class InvoiceDetailView(LoginRequiredMixin, DetailView):
    model = Invoice
    template_name = "invoice/detail.html"
    context_object_name = "invoice"

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Handle PDF viewing/downloading
        pdf_action = request.GET.get('pdf')
        if pdf_action in ('view', 'download'):
            from django.http import FileResponse, Http404, HttpResponse
            from django.conf import settings
            import os
            
            file_path = Path(settings.MEDIA_ROOT) / self.object.file_path
            if not file_path.exists():
                raise Http404("PDF файл не знайдено")
            
            filename = os.path.basename(self.object.file_path)
            
            # Read file content
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            # Create response
            if pdf_action == 'download':
                response = HttpResponse(file_content, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
            else:
                response = HttpResponse(file_content, content_type='application/pdf')
                response['Content-Disposition'] = f'inline; filename="{filename}"'
                # Allow embedding in iframe - override middleware X-Frame-Options
                response['X-Frame-Options'] = 'SAMEORIGIN'
            
            return response
        
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Sort items by item_id numerically (ascending)
        # Convert to list and sort by numeric value of item_id
        items_list = list(self.object.items.all())
        items_list.sort(key=lambda x: int(x.item_id) if x.item_id.isdigit() else float('inf'))
        
        # Decode Unicode escape sequences in item names for existing records
        # Create a wrapper to store decoded names without modifying the model
        class ItemWrapper:
            def __init__(self, item):
                self._item = item
                # Always decode Unicode escapes in item_name (even if not visible in DB)
                # This handles cases where text is stored correctly but displayed with escapes
                if item.item_name:
                    # Decode any Unicode escape sequences
                    decoded_name = decode_unicode_escapes(item.item_name)
                    # If decoding changed something, use decoded version
                    if decoded_name != item.item_name:
                        self.item_name = decoded_name
                    else:
                        self.item_name = item.item_name
                else:
                    self.item_name = item.item_name
                
            def __getattr__(self, name):
                return getattr(self._item, name)
        
        items_list = [ItemWrapper(item) for item in items_list]
        
        context['items'] = items_list
        context['form'] = InvoiceItemForm()
        
        # Get validation errors from invoice_data
        validation_errors = {}
        if self.object.invoice_data and isinstance(self.object.invoice_data, dict):
            validation_errors = self.object.invoice_data.get("validation_errors", {})
        context['validation_errors'] = validation_errors
        
        # Get car matches from invoice_data and convert UUID keys to strings for template
        car_matches = {}
        if self.object.invoice_data and isinstance(self.object.invoice_data, dict):
            raw_car_matches = self.object.invoice_data.get("car_matches", {})
            # Convert keys to strings if they're not already
            for key, value in raw_car_matches.items():
                car_matches[str(key)] = value
        context['car_matches'] = car_matches
        
        # Convert validation errors to JSON for JavaScript
        import json
        context['validation_errors_json'] = json.dumps(validation_errors)
        
        # Get total validation error and convert Decimal to float for display
        total_validation_error = None
        if self.object.invoice_data and isinstance(self.object.invoice_data, dict):
            total_validation_error = self.object.invoice_data.get("total_validation_error")
            if total_validation_error:
                # Convert Decimal values to float for template display
                if isinstance(total_validation_error, dict):
                    total_validation_error = {
                        k: float(v) if isinstance(v, (Decimal, int, float)) else v
                        for k, v in total_validation_error.items()
                    }
        context['total_validation_error'] = total_validation_error
        
        return context


class InvoiceItemUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            item = InvoiceItem.objects.get(uuid=pk)
        except InvoiceItem.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'errors': {'__all__': ['Позицію не знайдено']}
            }, status=404)

        form = InvoiceItemForm(request.POST, instance=item)
        
        if form.is_valid():
            form.save()
            return JsonResponse({
                'status': 'ok',
                'message': 'Позицію успішно оновлено'
            })
        
        return JsonResponse({
            'status': 'error',
            'errors': form.errors
        }, status=400)


class InvoiceItemDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            item = InvoiceItem.objects.get(uuid=pk)
            invoice_uuid = item.invoice.uuid
            item.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'ok',
                    'message': 'Позицію успішно видалено'
                })
            
            return redirect('invoice-detail', pk=invoice_uuid)
        except InvoiceItem.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'errors': {'__all__': ['Позицію не знайдено']}
                }, status=404)
            return redirect('invoice-list')


class InvoiceDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            invoice = Invoice.objects.get(uuid=pk)
            invoice.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'ok',
                    'message': 'Фактуру успішно видалено'
                })
            
            return redirect('invoice-list')
        except Invoice.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'errors': {'__all__': ['Фактуру не знайдено']}
                }, status=404)
            return redirect('invoice-list')
